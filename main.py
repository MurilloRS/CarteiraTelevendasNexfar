import os
import tkinter as tk
from tkinter import filedialog, scrolledtext

import openpyxl
import pandas as pd
import pyodbc

from sql import carteira, cliente_query, delete, insert, query, vendedor_query

# Construir a string de conexão
connection_string = (
    r'DRIVER={SQL Server};'
    r'SERVER=192.168.2.10;'
    r'DATABASE=MOINHO;'
    r'UID=sa;'
    r'PWD=moitgt2526;'
)

# Função para conectar ao banco de dados e verificar a existência da combinação (cd_clien, cd_vend)
def verificar_combinacao(cd_clien, cd_vend, aba):
    cd_clien_str = str(cd_clien)
    cd_vend_str = str(cd_vend)
    
    conn = pyodbc.connect(connection_string)
    cursor = conn.cursor()
    
    cursor.execute(query, (cd_clien_str, cd_vend_str))
    row = cursor.fetchall()
    
    cursor.execute(cliente_query, cd_clien)
    cliente_row = cursor.fetchall()
    
    if not cliente_row:
        text_box.insert(
            tk.END, f'\n(cd_clien = {cd_clien_str}, cd_vend = {cd_vend_str}) - cd_clien {cd_clien_str}: INVÁLIDO'
        )
        return
    
    cursor.execute(vendedor_query, cd_vend)
    vend_row = cursor.fetchall()
    
    if not vend_row:
        text_box.insert(
            tk.END, f'\n(cd_clien = {cd_clien_str}, cd_vend = {cd_vend_str}) - cd_vend {cd_vend_str}: INVÁLIDO'
        )
        return
    
    if row:
        if aba == "Remover":
            cursor.execute(delete, (cd_clien_str, cd_vend_str))
            conn.commit()
            text_box.insert(tk.END, f'\n(cd_clien = {cd_clien_str}, cd_vend = {cd_vend_str}) - REMOVIDO COM SUCESSO.')
        else:
            text_box.insert(tk.END, f'\n(cd_clien = {cd_clien_str}, cd_vend = {cd_vend_str}) - Já existe.')
    else:
        if aba == "Incluir":
            cursor.execute(insert, (cd_clien_str, cd_vend_str))
            conn.commit()
            text_box.insert(tk.END, f'\n(cd_clien = {cd_clien_str}, cd_vend = {cd_vend_str}) - INCLUÍDO COM SUCESSO.')
        else:
            text_box.insert(tk.END, f'\n(cd_clien = {cd_clien_str}, cd_vend = {cd_vend_str}) - Não existe.')


def extrair_carteira():
    conn = pyodbc.connect(connection_string)
    carteira_df = pd.read_sql_query(carteira, conn)
    conn.close()
    
    filename = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Planilha Excel", "*.xlsx")],
        initialfile="Carteira Televendas Nexfar.xlsx"
    )
    
    if filename:
        carteira_df.to_excel(filename, index=False)


def selecionar_planilha():
    text_box.delete('1.0', tk.END)
    
    filename = filedialog.askopenfilename(filetypes=[("Planilha Excel", "*.xlsx")])
    
    if filename:
        try:
            nome_arquivo = os.path.basename(filename)
            text_box.insert(tk.END, f"Arquivo: {nome_arquivo}")
            
            workbook = openpyxl.load_workbook(filename)
            
            for sheet_name in workbook.sheetnames:
                text_box.insert(tk.END, f"\n\nAba: {sheet_name}\n")
                sheet = workbook[sheet_name]
                
                expected_columns = ("cd_clien", "cd_vend")
                header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
                
                if header_row is None or not all(column_name in header_row for column_name in expected_columns):
                    raise ValueError("A planilha não possui as colunas esperadas.")
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):
                        verificar_combinacao(row[0], row[1], sheet_name)
            
            workbook.close()
        
        except FileNotFoundError:
            text_box.insert(tk.END, "Erro: Arquivo não encontrado.")
        except ValueError as e:
            text_box.insert(tk.END, f"Erro: {str(e)}")


# Criar a janela tkinter
root = tk.Tk()
root.title("Selecionar Planilha")

text_box = scrolledtext.ScrolledText(root, height=20, width=100)
text_box.pack(padx=20, pady=10, fill=tk.BOTH, expand=True)

button_selecionar_planilha = tk.Button(root, text="Selecionar Planilha", command=selecionar_planilha)
button_selecionar_planilha.pack(padx=20, pady=10)

button_extrair_carteira = tk.Button(root, text="Extrair Carteira", command=extrair_carteira)
button_extrair_carteira.pack(padx=20, pady=10, side=tk.LEFT)

root.mainloop()
